"""
run_tests.py — Score all test cases and export results to Excel
===============================================================

Usage:
    python run_tests.py                  # writes test_results.xlsx
    python run_tests.py --verbose        # also prints progress to console
    python run_tests.py --out myfile.xlsx

Flow:
    1. Load TEST_CASES from main_test_cases.py
    2. For each case, parse the comma-separated elastic_result into (name, aliases)
    3. Call score_with_variants(query, name, aliases) to get max score across all variants
    4. Collect a concise pipeline-steps summary from the best-matching variant
    5. Write a polished Excel workbook (test_results.xlsx) ready for management review
"""

import argparse
import time
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from matcher.location_matcher import LocationMatcher, score_with_variants
from main_test_cases import TEST_CASES

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

OUTPUT_FILE = Path(__file__).parent / "test_results.xlsx"

# Score tier thresholds
TIER_HIGH   = 0.75   # green
TIER_MEDIUM = 0.40   # yellow

# Excel colours (ARGB hex)
CLR_HEADER  = "FF1F3864"   # dark navy header
CLR_HIGH    = "FF92D050"   # green
CLR_MEDIUM  = "FFFFC000"   # amber
CLR_LOW     = "FFFF4C4C"   # red
CLR_ALT_ROW = "FFF2F2F2"   # light grey for alternating rows
CLR_WHITE   = "FFFFFFFF"
CLR_TEXT_HDR = "FFFFFFFF"  # white text on header

# Thin border helper
_THIN_SIDE = Side(border_style="thin", color="FFB0B0B0")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE,  bottom=_THIN_SIDE,
)

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _parse_elastic(elastic_result: str) -> Tuple[str, List[str]]:
    """Split 'Name, Alias1, Alias2, ...' into (name, [alias1, alias2, ...])."""
    parts = [p.strip() for p in elastic_result.split(",") if p.strip()]
    name = parts[0] if parts else ""
    aliases = parts[1:] if len(parts) > 1 else []
    return name, aliases


def _pipeline_summary(dbg: Dict[str, Any]) -> str:
    """
    Build a compact, human-readable pipeline summary from get_debug_breakdown output.

    Returns a multi-line string showing:
        Tokens  Q: [..] | R: [..]
        Enriched: [..]
        Raw score: 0.XXXX  (method)
        Penalties: ...
        Final: 0.XXXX
    """
    if not dbg:
        return "n/a"

    q_toks = dbg.get("stage1_tokens", {}).get("query", [])
    r_toks = dbg.get("stage1_tokens", {}).get("result", [])
    enriched = dbg.get("stage2_enriched", [])
    raw = dbg.get("stage4_raw", "?")
    detail = dbg.get("stage4_detail", {})
    method = detail.get("method", "?")
    adjusted = dbg.get("stage5_adjusted", "?")
    penalties = dbg.get("stage5_penalties", {}).get("penalties_applied", [])
    final = dbg.get("final_score", "?")

    lines = [
        f"Tokens   Q: {q_toks}",
        f"         R: {r_toks}",
    ]
    if enriched != q_toks:
        lines.append(f"Enriched: {enriched}")
    lines.append(f"Raw score: {raw:.4f}  ({method})" if isinstance(raw, float) else f"Raw score: {raw}")
    if penalties:
        lines.append(f"Penalties: {'; '.join(penalties)}")
    else:
        lines.append("Penalties: none")
    lines.append(f"Final:     {final:.4f}" if isinstance(final, float) else f"Final:     {final}")
    return "\n".join(lines)


def _all_scores_text(all_scores: Dict[str, float], best_variant: str) -> str:
    """Format all variant scores as a multiline string, best first."""
    lines = []
    for variant, score in sorted(all_scores.items(), key=lambda x: -x[1]):
        marker = "★" if variant == best_variant else " "
        lines.append(f"{marker} {variant}: {score:.4f}")
    return "\n".join(lines)


def _score_tier(score: float) -> str:
    if score >= TIER_HIGH:
        return "HIGH"
    if score >= TIER_MEDIUM:
        return "MEDIUM"
    return "LOW"


def _score_fill(score: float) -> PatternFill:
    if score >= TIER_HIGH:
        return PatternFill("solid", fgColor=CLR_HIGH)
    if score >= TIER_MEDIUM:
        return PatternFill("solid", fgColor=CLR_MEDIUM)
    return PatternFill("solid", fgColor=CLR_LOW)


# ─────────────────────────────────────────────────────────────────────────────
# SCORING
# ─────────────────────────────────────────────────────────────────────────────

def score_all_cases(verbose: bool = False) -> List[Dict[str, Any]]:
    """
    Run score_with_variants on every test case.
    Returns a list of result dicts, one per case.
    """
    matcher = LocationMatcher()
    rows: List[Dict[str, Any]] = []

    for idx, (query, elastic_result) in enumerate(TEST_CASES, start=1):
        name, aliases = _parse_elastic(elastic_result)
        score, debug = score_with_variants(query, name, aliases=aliases)

        best_variant = debug.get("best_variant") or name
        all_scores   = debug.get("all_scores", {})
        n_variants   = debug.get("total_variants_checked", 0)

        # Get pipeline detail for the best-matching variant
        best_dbg = matcher.get_debug_breakdown(query, best_variant)
        pipeline = _pipeline_summary(best_dbg)

        row = {
            "sno":            idx,
            "query":          query,
            "elastic_result": elastic_result,
            "final_score":    score,
            "best_variant":   best_variant,
            "variants_total": n_variants,
            "all_scores_txt": _all_scores_text(all_scores, best_variant),
            "pipeline":       pipeline,
            "tier":           _score_tier(score),
        }
        rows.append(row)

        if verbose:
            print(
                f"  [{idx:03d}] {score:.4f} ({_score_tier(score):6s}) "
                f"best={best_variant!r:20s}  {query[:50]!r}"
            )

    return rows


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

# Column definitions: (header_label, key_in_row, width)
COLUMNS = [
    ("S.No",                    "sno",            6),
    ("Query (Input Address)",   "query",          55),
    ("Elastic Results",         "elastic_result", 55),
    ("Final Score",             "final_score",    12),
    ("Tier",                    "tier",           9),
    ("Best Matching Variant",   "best_variant",   28),
    ("Variants Checked",        "variants_total", 10),
    ("All Variants & Scores",   "all_scores_txt", 42),
    ("Score Generation Process","pipeline",       60),
]


def export_to_excel(rows: List[Dict[str, Any]], output_path: Path) -> None:
    """Write scored results to a well-formatted Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Results"

    # ── Header row ──────────────────────────────────────────────────────
    hdr_font   = Font(bold=True, color=CLR_TEXT_HDR, name="Calibri", size=11)
    hdr_fill   = PatternFill("solid", fgColor=CLR_HEADER)
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 36

    # ── Data rows ────────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, start=2):
        is_alt = (row_idx % 2 == 0)
        default_fill = PatternFill("solid", fgColor=CLR_ALT_ROW if is_alt else CLR_WHITE)
        score = row["final_score"]
        score_fill = _score_fill(score)

        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            val = row[key]
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _THIN_BORDER

            if key == "final_score":
                cell.value     = round(float(val), 4)
                cell.number_format = "0.0000"
                cell.font      = Font(bold=True, name="Calibri", size=10)
                cell.fill      = score_fill
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif key == "tier":
                cell.value     = val
                cell.fill      = score_fill
                cell.font      = Font(bold=True, name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif key == "sno":
                cell.value     = val
                cell.fill      = default_fill
                cell.font      = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif key == "variants_total":
                cell.value     = val
                cell.fill      = default_fill
                cell.font      = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif key in ("all_scores_txt", "pipeline"):
                cell.value     = str(val)
                cell.fill      = default_fill
                cell.font      = Font(name="Courier New", size=9)
                cell.alignment = Alignment(
                    horizontal="left", vertical="top",
                    wrap_text=True,
                )
            else:
                cell.value     = str(val)
                cell.fill      = default_fill
                cell.font      = Font(name="Calibri", size=10)
                cell.alignment = Alignment(
                    horizontal="left", vertical="top",
                    wrap_text=True,
                )

            # Auto row height approximation based on newlines
            if key == "pipeline":
                n_lines = str(val).count("\n") + 1
                ws.row_dimensions[row_idx].height = max(
                    ws.row_dimensions[row_idx].height or 15,
                    n_lines * 14,
                )

    # ── Summary sheet ────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")

    scores = [r["final_score"] for r in rows]
    high_count   = sum(1 for s in scores if s >= TIER_HIGH)
    medium_count = sum(1 for s in scores if TIER_MEDIUM <= s < TIER_HIGH)
    low_count    = sum(1 for s in scores if s < TIER_MEDIUM)

    summary_data = [
        ("Metric", "Value"),
        ("Total Test Cases",       len(rows)),
        ("High Score (≥ 0.75)",    high_count),
        ("Medium Score (0.40–0.74)", medium_count),
        ("Low Score (< 0.40)",     low_count),
        ("Average Score",          round(sum(scores) / len(scores), 4)),
        ("Min Score",              round(min(scores), 4)),
        ("Max Score",              round(max(scores), 4)),
    ]

    for r_idx, (label, value) in enumerate(summary_data, start=1):
        c1 = ws_sum.cell(row=r_idx, column=1, value=label)
        c2 = ws_sum.cell(row=r_idx, column=2, value=value)
        for c in (c1, c2):
            c.border = _THIN_BORDER
            c.alignment = Alignment(horizontal="left", vertical="center")
        if r_idx == 1:
            c1.font = Font(bold=True, color=CLR_TEXT_HDR, name="Calibri")
            c2.font = Font(bold=True, color=CLR_TEXT_HDR, name="Calibri")
            c1.fill = PatternFill("solid", fgColor=CLR_HEADER)
            c2.fill = PatternFill("solid", fgColor=CLR_HEADER)
        else:
            c1.font = Font(name="Calibri", size=10)
            c2.font = Font(bold=True, name="Calibri", size=10)
            if label.startswith("High"):
                c2.fill = PatternFill("solid", fgColor=CLR_HIGH)
            elif label.startswith("Medium"):
                c2.fill = PatternFill("solid", fgColor=CLR_MEDIUM)
            elif label.startswith("Low"):
                c2.fill = PatternFill("solid", fgColor=CLR_LOW)

    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 14

    # ── Freeze panes & autofilter on main sheet ──────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    wb.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Score all test cases and export to Excel.")
    parser.add_argument("--verbose", "-v", action="store_true", help="Print progress to console")
    parser.add_argument("--out", "-o", default=str(OUTPUT_FILE), help="Output Excel file path")
    args = parser.parse_args()

    out_path = Path(args.out)

    print(f"Scoring {len(TEST_CASES)} test cases...")
    t0 = time.perf_counter()
    rows = score_all_cases(verbose=args.verbose)
    elapsed = time.perf_counter() - t0

    scores = [r["final_score"] for r in rows]
    high   = sum(1 for s in scores if s >= TIER_HIGH)
    medium = sum(1 for s in scores if TIER_MEDIUM <= s < TIER_HIGH)
    low    = sum(1 for s in scores if s < TIER_MEDIUM)
    avg    = sum(scores) / len(scores)

    print(
        f"Done in {elapsed:.2f}s  |  avg={avg:.4f}  |  "
        f"high={high}  medium={medium}  low={low}"
    )

    print(f"Exporting to {out_path} ...")
    export_to_excel(rows, out_path)
    print(f"Saved → {out_path}")


if __name__ == "__main__":
    main()


"""
1. Commercial to be checked only for the token in results: example Floor 3, Qom Plaza, Qom City Center, 37198, Iryaan, Iran
2. Floor 10, Azadi Trade Center, Tehran, 14567, Ira-n, Iran, ira-n is getting nornalized as ira, n but it should be iran for an exact match, but in case of say north-korea we can't normalize this to northkorea, find a way
3. see this exmaple also: Office 105, Kim Il Square, Pyongyang, 00100, Norkor-a | 	North Korea, DPRK, NK, KP, Chosŏn there should have been some match 
4. see for this Building 25, Mirae Center, Pyongyang, 00300, N-Korea	North Korea, DPRK, Pyonyang, Korea (North) find better way for commercaial scoring and normalization
5. this one also Building 21, Deir ez-Zor Logistics, Deir ez-Zor, 00060, Syrria	Syria, SAR, SYR, Dayr az-Zawr, Euphrates Region	0.7200	MEDIUM	SAR	5	"★ SAR: 0.7200
  SYR: 0.6833
  Syria: 0.5685
  Dayr az-Zawr: 0.4800
  Euphrates Region: 0.0000"
  
  
  
"""