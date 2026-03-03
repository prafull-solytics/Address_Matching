"""
generate_test_tables.py
=======================
Parses test_edge_cases.py and extracts every test case into:
  - test_cases.xlsx   (Excel workbook with one sheet per EC class)
  - test_cases.md     (Markdown tables, one section per EC class)

Columns: #, Test Method, Query, Elastic Result, Expected Score / Condition
"""

import ast
import re
import textwrap
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── paths ─────────────────────────────────────────────────────────────────
ROOT = Path(__file__).parent
SRC  = ROOT / "test_edge_cases.py"
XLSX = ROOT / "test_cases.xlsx"
MD   = ROOT / "test_cases.md"

# ── helpers ───────────────────────────────────────────────────────────────

def _lit(node) -> str:
    """Return a readable string for a constant / joined-string AST node."""
    if node is None:
        return ""
    if isinstance(node, ast.Constant):
        return repr(node.value) if isinstance(node.value, str) else str(node.value)
    if isinstance(node, ast.JoinedStr):          # f-string
        return "<f-string>"
    if isinstance(node, ast.Name):
        return node.id
    if isinstance(node, ast.Attribute):
        return f"{_lit(node.value)}.{node.attr}"
    return ast.unparse(node)


def _str_val(node) -> str:
    """Return the plain string value (no quotes) for a Constant str node."""
    if isinstance(node, ast.Constant) and isinstance(node.value, str):
        return node.value
    return _lit(node)


def _format_score(lo, hi) -> str:
    if lo == hi:
        return str(lo)
    if lo == 0.0 and hi == 0.0:
        return "0.0"
    return f"[{lo}, {hi}]"


# ── AST walking helpers ────────────────────────────────────────────────────

def get_calls(body: list, func_name: str) -> list[ast.Call]:
    """Collect all Call nodes whose function name ends with func_name."""
    results = []
    for node in ast.walk(ast.Module(body=body, type_ignores=[])):
        if isinstance(node, ast.Call):
            fn = node.func
            if isinstance(fn, ast.Attribute) and fn.attr == func_name:
                results.append(node)
            elif isinstance(fn, ast.Name) and fn.id == func_name:
                results.append(node)
    return results


def first_assert_score(body) -> tuple[str, str] | None:
    """Return (lo, hi) from the first assert_score call, or None."""
    calls = get_calls(body, "assert_score")
    if calls:
        c = calls[0]
        args = c.args
        if len(args) >= 3:
            lo = ast.literal_eval(args[1])
            hi = ast.literal_eval(args[2])
            return lo, hi
    return None


def first_match_call(body) -> tuple[str, str] | None:
    """Return (query, result) from the first matcher.match() call."""
    calls = get_calls(body, "match")
    for c in calls:
        # filter only matcher.match(...)
        if isinstance(c.func, ast.Attribute) and c.func.attr == "match":
            if len(c.args) >= 2:
                return _str_val(c.args[0]), _str_val(c.args[1])
    return None


def extract_expected(body) -> str:
    """
    Try to determine the expected score / condition from a test function body.
    Priority:
      1. assert_score → [lo, hi]
      2. == 0.0 / == 1.0 assert
      3. pytest.raises
      4. comparative assert (>, >=, <)
      5. assert score >= / <=
    """
    # 1. assert_score
    calls = get_calls(body, "assert_score")
    if calls:
        c = calls[0]
        args = c.args
        if len(args) >= 3:
            lo = ast.literal_eval(args[1])
            hi = ast.literal_eval(args[2])
            return _format_score(lo, hi)

    # 2. pytest.raises
    for node in ast.walk(ast.Module(body=body, type_ignores=[])):
        if isinstance(node, ast.Call):
            fn = node.func
            if isinstance(fn, ast.Attribute) and fn.attr == "raises":
                exc = node.args[0] if node.args else None
                return f"raises {_lit(exc)}"

    # Walk all Assert nodes for inline patterns
    for node in ast.walk(ast.Module(body=body, type_ignores=[])):
        if not isinstance(node, ast.Assert):
            continue
        test = node.test

        # == 0.0 / == 1.0 / == exact
        if isinstance(test, ast.Compare):
            if len(test.ops) == 1 and isinstance(test.ops[0], ast.Eq):
                right = test.comparators[0]
                if isinstance(right, ast.Constant):
                    return f"== {right.value}"
            # >= threshold
            if len(test.ops) == 1 and isinstance(test.ops[0], ast.GtE):
                right = test.comparators[0]
                if isinstance(right, ast.Constant):
                    return f">= {right.value}"
            # <= threshold
            if len(test.ops) == 1 and isinstance(test.ops[0], ast.LtE):
                right = test.comparators[0]
                if isinstance(right, ast.Constant):
                    return f"<= {right.value}"

        # a > b  (comparative)
        if isinstance(test, ast.Compare):
            if len(test.ops) == 1 and isinstance(test.ops[0], ast.Gt):
                return f"{ast.unparse(test.left)} > {ast.unparse(test.comparators[0])}"
            if len(test.ops) == 1 and isinstance(test.ops[0], ast.GtE):
                return f"{ast.unparse(test.left)} >= {ast.unparse(test.comparators[0])}"

    return "see test"


def match_calls_from_body(body) -> list[tuple[str, str]]:
    """Return all (query, result) pairs from matcher.match() in the body."""
    pairs = []
    calls = get_calls(body, "match")
    for c in calls:
        if isinstance(c.func, ast.Attribute) and c.func.attr == "match":
            if len(c.args) >= 2:
                pairs.append((_str_val(c.args[0]), _str_val(c.args[1])))
    return pairs


# ── main extractor ────────────────────────────────────────────────────────

class TestCase:
    __slots__ = ("ec_id", "class_name", "method", "query", "result", "expected", "notes")

    def __init__(self, ec_id, class_name, method, query, result, expected, notes=""):
        self.ec_id      = ec_id
        self.class_name = class_name
        self.method     = method
        self.query      = query
        self.result     = result
        self.expected   = expected
        self.notes      = notes


def extract_parametrize_args(decorator) -> list[tuple[str, str, str]] | None:
    """
    Parse @pytest.mark.parametrize("query,result,label", [...])
    Returns list of (query, result, label).
    """
    if not isinstance(decorator, ast.Call):
        return None
    fn = decorator.func
    if not (isinstance(fn, ast.Attribute) and fn.attr == "parametrize"):
        return None
    if len(decorator.args) < 2:
        return None

    # First arg: parameter names string
    names_node = decorator.args[0]
    if not isinstance(names_node, ast.Constant):
        return None
    names = [n.strip() for n in names_node.value.split(",")]

    # Second arg: list of tuples
    values_node = decorator.args[1]
    if not isinstance(values_node, (ast.List, ast.Tuple)):
        return None

    rows = []
    for elt in values_node.elts:
        if isinstance(elt, ast.Tuple) and len(elt.elts) == len(names):
            vals = [_str_val(e) for e in elt.elts]
            # map by name index
            q = r = label = ""
            for i, name in enumerate(names):
                if name in ("query", "q"):
                    q = vals[i]
                elif name in ("result", "r"):
                    r = vals[i]
                elif name in ("label",):
                    label = vals[i]
            rows.append((q, r, label))
    return rows if rows else None


def parse_test_file(path: Path) -> list[TestCase]:
    source = path.read_text(encoding="utf-8")
    tree = ast.parse(source)

    # Class-level constant strings (for LONG_US etc.)
    class_attrs: dict[str, dict[str, str]] = {}

    cases: list[TestCase] = []

    for node in tree.body:
        if not isinstance(node, ast.ClassDef):
            continue

        cls_name = node.name
        # ec_id from "TestEC1_..." → "EC-1"
        m = re.match(r"TestEC(\d+)_", cls_name)
        ec_id = f"EC-{m.group(1)}" if m else cls_name

        # Collect class-level string attributes
        cls_consts: dict[str, str] = {}
        for item in node.body:
            if isinstance(item, ast.Assign):
                for target in item.targets:
                    if isinstance(target, ast.Name) and isinstance(item.value, ast.Constant):
                        cls_consts[target.id] = str(item.value.value)
        class_attrs[cls_name] = cls_consts

        for item in node.body:
            if not isinstance(item, ast.FunctionDef):
                continue
            if not item.name.startswith("test_"):
                continue

            method = item.name

            # ── check for @pytest.mark.parametrize ─────────────────────
            param_rows = None
            expected_for_param = "see test"
            for dec in item.decorator_list:
                rows = extract_parametrize_args(dec)
                if rows:
                    param_rows = rows
                    # Try to infer expected from body
                    expected_for_param = extract_expected(item.body)
                    break

            if param_rows:
                for q, r, label in param_rows:
                    cases.append(TestCase(
                        ec_id, cls_name, method,
                        q, r, expected_for_param,
                        notes=label,
                    ))
                continue

            # ── resolve self.LONG_XX references ────────────────────────
            # replace Attribute accesses like self.LONG_US in match calls
            body_src = ast.unparse(ast.Module(body=item.body, type_ignores=[]))

            # ── get all match calls ─────────────────────────────────────
            pairs = match_calls_from_body(item.body)

            # Resolve class attribute references (e.g. self.LONG_US)
            resolved_pairs = []
            for q, r in pairs:
                # If q/r look like attribute refs "self.LONG_US"
                # ast.unparse gives "self.LONG_US" → look up in cls_consts
                for attr_name, attr_val in cls_consts.items():
                    q = q.replace(f"self.{attr_name}", attr_val)
                    r = r.replace(f"self.{attr_name}", attr_val)
                resolved_pairs.append((q, r))

            expected = extract_expected(item.body)

            # Check for pytest.raises (TypeError tests)
            raises_check = False
            for n in ast.walk(ast.Module(body=item.body, type_ignores=[])):
                if isinstance(n, ast.Call):
                    fn2 = n.func
                    if isinstance(fn2, ast.Attribute) and fn2.attr == "raises":
                        raises_check = True
                        exc_name = _lit(n.args[0]) if n.args else "Exception"
                        if not resolved_pairs:
                            resolved_pairs = [("(invalid input)", "(invalid input)")]
                        expected = f"raises {exc_name}"
                        break

            if not resolved_pairs:
                # No match call found – note it but still record
                cases.append(TestCase(ec_id, cls_name, method, "—", "—", expected))
            else:
                for q, r in resolved_pairs:
                    cases.append(TestCase(ec_id, cls_name, method, q, r, expected))

    return cases


# ── Excel writer ──────────────────────────────────────────────────────────

HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
EC_FILL    = PatternFill("solid", fgColor="D6E4F0")
EC_FONT    = Font(bold=True, color="1F4E79", size=10)
EVEN_FILL  = PatternFill("solid", fgColor="F2F8FF")
ODD_FILL   = PatternFill("solid", fgColor="FFFFFF")
CELL_FONT  = Font(size=9)
CELL_ALIGN = Alignment(wrap_text=True, vertical="top")
THIN       = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = ["#", "EC", "Class", "Test Method", "Query", "Elastic Result", "Expected Score / Condition", "Notes"]
COL_WIDTHS = [5, 7, 32, 38, 45, 30, 35, 40]


def write_excel(cases: list[TestCase], path: Path):
    wb = openpyxl.Workbook()

    # ── ALL sheet ──────────────────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = "All Test Cases"
    _write_sheet(ws_all, cases)

    # ── Per-EC sheets ──────────────────────────────────────────────────
    ec_ids = list(dict.fromkeys(c.ec_id for c in cases))
    for ec in ec_ids:
        subset = [c for c in cases if c.ec_id == ec]
        ws = wb.create_sheet(title=ec)
        _write_sheet(ws, subset)

    wb.save(path)
    print(f"✅  Saved Excel  → {path}")


def _write_sheet(ws, cases: list[TestCase]):
    # Header row
    for col_idx, (hdr, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for row_idx, tc in enumerate(cases, start=2):
        row_fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL
        row_data = [
            row_idx - 1,
            tc.ec_id,
            tc.class_name,
            tc.method,
            tc.query,
            tc.result,
            tc.expected,
            tc.notes,
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = CELL_FONT
            cell.fill      = row_fill
            cell.alignment = CELL_ALIGN
            cell.border    = THIN_BORDER
        ws.row_dimensions[row_idx].height = 30


# ── Markdown writer ───────────────────────────────────────────────────────

def md_escape(s: str) -> str:
    """Escape pipe and newline characters for Markdown tables."""
    return str(s).replace("|", "\\|").replace("\n", " ").replace("\r", "")


def write_markdown(cases: list[TestCase], path: Path):
    lines = ["# LocationMatcher — Test Cases\n",
             f"> Auto-generated from `test_edge_cases.py`  \n",
             f"> Total test cases: **{len(cases)}**\n",
             "---\n"]

    ec_ids = list(dict.fromkeys(c.ec_id for c in cases))

    for ec in ec_ids:
        subset = [c for c in cases if c.ec_id == ec]
        # Section title from the first case's class name
        cls_raw = subset[0].class_name  # e.g. TestEC1_AbbreviationAmbiguity
        title_m = re.match(r"TestEC\d+_(.*)", cls_raw)
        title = title_m.group(1).replace("_", " ") if title_m else cls_raw
        lines.append(f"\n## {ec} — {title}\n")
        lines.append(f"*{len(subset)} test case(s)*\n")

        # Table header
        lines.append("| # | Test Method | Query | Elastic Result | Expected Score / Condition | Notes |")
        lines.append("|---|-------------|-------|----------------|---------------------------|-------|")

        for i, tc in enumerate(subset, start=1):
            row = (
                f"| {i} "
                f"| `{md_escape(tc.method)}` "
                f"| `{md_escape(tc.query)}` "
                f"| `{md_escape(tc.result)}` "
                f"| {md_escape(tc.expected)} "
                f"| {md_escape(tc.notes)} |"
            )
            lines.append(row)

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"✅  Saved Markdown → {path}")


# ── entry point ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"Parsing {SRC} …")
    cases = parse_test_file(SRC)
    print(f"  Found {len(cases)} test entries across {len(set(c.ec_id for c in cases))} EC groups")

    write_excel(cases, XLSX)
    write_markdown(cases, MD)

    print("\nDone! Files written to project root:")
    print(f"  {XLSX.name}")
    print(f"  {MD.name}")

